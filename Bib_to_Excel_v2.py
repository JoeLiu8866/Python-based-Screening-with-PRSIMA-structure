import bibtexparser
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from copy import copy
import os

def fill_excel_with_data(template_path, data_path, encoding='utf-8'):
    # Parse BibTeX file
    try:
        with open(data_path, 'r', encoding=encoding) as bibtex_file:
            bib_database = bibtexparser.load(bibtex_file)
    except UnicodeDecodeError:
        print(f"Unable to read the file using {encoding} encoding. Please try a different encoding format.")
        return False

    # Extract required fields
    data = []
    for entry in bib_database.entries:
        data.append({
            "Title": entry.get("title", ""),
            "href": entry.get("url", ""),
            "year": entry.get("year", ""),
            "Venue": entry.get("booktitle", entry.get("journal", "")),
            "Author": entry.get("author", ""),
            "Abstract": entry.get("abstract", "")
        })

    # Convert to DataFrame
    df = pd.DataFrame(data)

    # Load Excel template
    try:
        wb = load_workbook(template_path)
        ws = wb.active
    except Exception as e:
        print(f"Unable to open the Excel template file. Please check if the file path and format are correct: {e}")
        return False

    # Define column mapping, columns start from 1
    column_mapping = {
        "Title": 1,
        "href": 2,
        "year": 3,
        "Venue": 4,
        "Author": 5,
        "Abstract": 6
    }

    # Get header and content row styles
    header_styles = {}
    content_styles = {}

    for col in column_mapping.values():
        header_styles[col] = {
            'font': ws.cell(row=3, column=col).font,
            'border': ws.cell(row=3, column=col).border,
            'fill': ws.cell(row=3, column=col).fill,
            'number_format': ws.cell(row=3, column=col).number_format,
            'protection': ws.cell(row=3, column=col).protection,
            'alignment': ws.cell(row=3, column=col).alignment
        }
        content_styles[col] = {
            'font': ws.cell(row=4, column=col).font,
            'border': ws.cell(row=4, column=col).border,
            'fill': ws.cell(row=4, column=col).fill,
            'number_format': ws.cell(row=4, column=col).number_format,
            'protection': ws.cell(row=4, column=col).protection,
            'alignment': ws.cell(row=4, column=col).alignment
        }

    # Fill data starting from row 4
    start_row = 4

    for index, row in df.iterrows():
        for field, col in column_mapping.items():
            cell = ws.cell(row=start_row + index, column=col)
            if field == "year":
                cell.value = int(row[field]) if pd.notnull(row[field]) else None
            else:
                cell.value = row[field]
            cell.alignment = Alignment(wrap_text=True)

    # Restore content row styles and enable word wrap, font size 8
    for row in range(start_row, start_row + len(df)):
        for col in column_mapping.values():
            cell = ws.cell(row=row, column=col)
            style = content_styles[col]
            cell.font = Font(size=8, name=style['font'].name, bold=style['font'].bold, italic=style['font'].italic, vertAlign=style['font'].vertAlign, underline=style['font'].underline, strike=style['font'].strike, color=style['font'].color)
            cell.border = copy(style['border'])
            cell.fill = copy(style['fill'])
            cell.number_format = copy(style['number_format'])
            cell.protection = copy(style['protection'])
            cell.alignment = Alignment(
                wrap_text=True,
                horizontal=style['alignment'].horizontal,
                vertical=style['alignment'].vertical
            )
            if col == 3:  # Third column (year column)
                cell.number_format = '0'

    # Restore header row styles
    for row in range(1, 4):
        for col in column_mapping.values():
            cell = ws.cell(row=row, column=col)
            style = header_styles[col]
            cell.font = copy(style['font'])
            cell.border = copy(style['border'])
            cell.fill = copy(style['fill'])
            cell.number_format = copy(style['number_format'])
            cell.protection = copy(style['protection'])
            cell.alignment = copy(style['alignment'])

    # Set row height: content rows 100, header rows 20
    for row in range(4, ws.max_row + 1):
        ws.row_dimensions[row].height = 100

    for row in range(1, 4):
        ws.row_dimensions[row].height = 20

    # Extract BibTeX filename and generate Excel filename
    bibtex_filename = os.path.splitext(os.path.basename(data_path))[0]
    output_folder = os.path.dirname(data_path)
    output_excel_path = os.path.join(output_folder, f"{bibtex_filename}.xlsx")

    # Save new Excel file
    try:
        wb.save(output_excel_path)
        print(f"File saved to: {output_excel_path}")
        return True
    except PermissionError:
        print(f"Permission denied: unable to save file to {output_excel_path}. Please check if you have write permissions or if the file is being used by another program.")
        return False

def process_bib_files_in_directory(template_path, root_directory, encoding='utf-8'):
    bib_files = []
    for root, dirs, files in os.walk(root_directory):
        for file in files:
            if file.endswith('.bib'):
                bib_files.append(os.path.join(root, file))
    
    total_files = len(bib_files)
    processed_files = 0

    for bib_file_path in bib_files:
        success = fill_excel_with_data(template_path, bib_file_path, encoding)
        if success:
            processed_files += 1
    
    print(f"Found {total_files} BibTeX files in the directory.")
    print(f"Successfully processed {processed_files} files.")
    if total_files != processed_files:
        print(f"{total_files - processed_files} files failed to process.")

if __name__ == "__main__":
    template_path = input("Please enter the path to the Excel template file: ").strip().strip('"')
    root_directory = input("Please enter the root directory containing the BibTeX files: ").strip().strip('"')
    encoding = input("Please enter the file encoding format (default utf-8): ").strip() or 'utf-8'
    
    process_bib_files_in_directory(template_path, root_directory, encoding)

